import openpyxl
wb=openpyxl.load_workbook("D:\\Study material\\NIIT\\Pythonfile\\Excel Using Python\\Demo.xlsx")

sh1=wb['Name']

rows=sh1.max_row
column=sh1.max_column
copy=[]


from openpyxl import Workbook

wb=Workbook()
wb['Sheet'].title="Report"
sh2=wb.active

start=1
for i in range(1,rows+1):
    if sh1.cell(i,1).value not in copy:
        copy.append(sh1.cell(i,1).value)
        for j in range(1,column+1):
            sh2.cell(start,j,sh1.cell(i,j).value)
        start+=1



wb.save("Finalreport2.xlsx")